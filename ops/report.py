"""
Excel reporting for Use Case 2 - the Common Scope's "Dashboards &
Reporting: common reporting layer covering document status, processing
volumes, accuracy, and exceptions".

Four sheets:
  Documents    every processed document with the full metadata table,
               classification confidence, routing, and transaction key
  Transactions one row per correlated transaction
  Audit Packs  pack composition per transaction - present / missing /
               optional-not-provided per item
  Exceptions   the human-in-the-loop queue: every review flag raised
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logger import get_logger
from ops.models import METADATA_FIELDS, AuditPack, OpsDocument
from ops.ops_config import OPS_REPORT_NAME, OPS_REPORTS_DIR, ensure_output_dirs

logger = get_logger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
FILL_COMPLETE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_INCOMPLETE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_REVIEW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: list[str], max_width: int = 42) -> None:
    for col, header in enumerate(headers, start=1):
        longest = max(
            [len(str(header))]
            + [len(str(ws.cell(row=r, column=col).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col)].width = min(longest + 2, max_width)


def build_ops_report(documents: list[OpsDocument], packs: list[AuditPack]) -> Path:
    ensure_output_dirs()
    wb = Workbook()

    ws = wb.active
    ws.title = "Documents"
    doc_headers = (
        ["Source File", "Document Type", "Confidence", "Assigned Team", "Transaction Key"]
        + [f.replace("_", " ").title() for f in METADATA_FIELDS]
        + ["Review Flags", "Filed Path"]
    )
    _write_header(ws, doc_headers)
    for doc in documents:
        ws.append(
            [doc.source_file, doc.doc_type_name, f"{doc.classification_confidence:.0f}%",
             doc.assigned_team, doc.transaction_key]
            + [doc.meta(f) for f in METADATA_FIELDS]
            + ["; ".join(doc.review_flags), doc.filed_path]
        )
        if doc.review_flags:
            for col in range(1, len(doc_headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = FILL_REVIEW
    _autosize(ws, doc_headers)

    ws_t = wb.create_sheet("Transactions")
    t_headers = ["Transaction Key", "Type", "Portfolio Code", "Portfolio Name", "Client",
                 "Date", "Amount", "Trade ID", "Salesperson", "Documents", "Pack Status", "Audit Folder"]
    _write_header(ws_t, t_headers)
    for pack in packs:
        tx = pack.transaction
        ws_t.append([tx.transaction_key, tx.transaction_type, tx.portfolio_code,
                     tx.portfolio_name, tx.client_name, tx.transaction_date,
                     tx.transaction_amount, tx.trade_id, tx.salesperson, len(tx.documents),
                     pack.status, pack.audit_folder])
        fill = FILL_COMPLETE if pack.is_complete else FILL_INCOMPLETE
        for col in range(1, len(t_headers) + 1):
            ws_t.cell(row=ws_t.max_row, column=col).fill = fill
    _autosize(ws_t, t_headers)

    ws_p = wb.create_sheet("Audit Packs")
    p_headers = ["Transaction Key", "Type", "Salesperson", "Pack Item", "Required", "Present", "Satisfied By"]
    _write_header(ws_p, p_headers)
    for pack in packs:
        for item in pack.items:
            ws_p.append([pack.transaction.transaction_key, pack.transaction.transaction_type,
                         pack.transaction.salesperson, item.name,
                         "Yes" if item.required else "Where applicable",
                         "Yes" if item.present else "No", item.satisfied_by_code])
            if item.required and not item.present:
                for col in range(1, len(p_headers) + 1):
                    ws_p.cell(row=ws_p.max_row, column=col).fill = FILL_INCOMPLETE
    _autosize(ws_p, p_headers)

    ws_e = wb.create_sheet("Exceptions")
    e_headers = ["Source File", "Document Type", "Transaction Key", "Review Flag"]
    _write_header(ws_e, e_headers)
    for doc in documents:
        for flag in doc.review_flags:
            ws_e.append([doc.source_file, doc.doc_type_name, doc.transaction_key, flag])
    _autosize(ws_e, e_headers)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OPS_REPORTS_DIR / OPS_REPORT_NAME.replace(".xlsx", f"_{stamp}.xlsx")
    wb.save(out)
    logger.info("Saved Ops report to %s", out)
    return out
