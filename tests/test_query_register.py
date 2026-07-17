from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models.schemas import (
    ExtractionResult,
    FieldValue,
    PreValidationFlag,
    ProcessingLogEntry,
    ValidationReport,
    ValidationStatus,
    FieldValidationResult,
)
from app.services.query_register import QueryRegisterBuilder, recon_type_for_form_code


def _make_extraction(form_code: str, full_name: str = "Jane Doe") -> ExtractionResult:
    return ExtractionResult(
        source_file="test.pdf",
        form_code=form_code,
        fields={"full_name": FieldValue(field_id="full_name", value=full_name, confidence=95.0)},
    )


def _make_log_entry(instruction_status: str, rejection_reason: str = "") -> ProcessingLogEntry:
    return ProcessingLogEntry.now(
        original_filename="test.pdf",
        new_filename="ADD_E1_DOE_20260707.pdf",
        form_type_detected="ADD",
        classification_confidence=95.0,
        validation_status="PASS",
        destination_path="/tmp/out",
        instruction_status=instruction_status,
        rejection_reason=rejection_reason,
        channel="Email",
        upload_date="2026-07-07T09:00:00",
    )


def test_recon_type_mapping():
    assert recon_type_for_form_code("ADD") == "Additional Investments"
    assert recon_type_for_form_code("DIS") == "Withdrawals"
    assert recon_type_for_form_code("DIS_GSG") == "Withdrawals"
    assert recon_type_for_form_code("KYC") is None
    assert recon_type_for_form_code("UNKNOWN_CODE") is None


def test_clean_instruction_not_logged_but_counted():
    qr = QueryRegisterBuilder()
    extraction = _make_extraction("ADD")
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[
        FieldValidationResult(field_id="full_name", value="Jane Doe", status=ValidationStatus.PASS),
    ])
    log_entry = _make_log_entry("Captured")
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[])

    assert len(qr._query_rows) == 0
    assert qr._recon_counts["Additional Investments"]["Received"] == 1
    assert qr._recon_counts["Additional Investments"]["Processed"] == 1


def test_rejected_instruction_logged():
    qr = QueryRegisterBuilder()
    extraction = _make_extraction("DIS")
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
    log_entry = _make_log_entry("Rejected", rejection_reason="wrong banking details")
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[])

    assert len(qr._query_rows) == 1
    row, risk = qr._query_rows[0]
    # Enquiry type now comes from the form code (DIS -> Withdrawal instruction),
    # matching the updated template's instruction-oriented picklist.
    assert row["Type of Enquiry"] == "Instruction-Withdrawal"
    # The problem goes to Sales Comments To Ops; Query Description describes
    # the instruction itself (falls back to a plain label here since no
    # amount/fund was extracted).
    assert "wrong banking details" in row["Sales Comments To Ops"]
    assert row["Query Description"] == "Withdrawal"
    assert risk == "High"
    assert qr._recon_counts["Withdrawals"]["Pending"] == 1


def test_triggered_flag_logged_even_when_captured():
    qr = QueryRegisterBuilder()
    extraction = _make_extraction("STATIC")
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
    log_entry = _make_log_entry("Captured")
    flag = PreValidationFlag(
        flag_id="kyc_missing", label="KYC companion document missing",
        triggered=True, rejection_risk="Medium", reason="No KYC form found in batch",
    )
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[flag])

    assert len(qr._query_rows) == 1
    row, risk = qr._query_rows[0]
    # STATIC form -> "Instruction-Static" enquiry tag; the flag detail lives
    # in Sales Comments To Ops, not the Query Description.
    assert row["Type of Enquiry"] == "Instruction-Static"
    assert "KYC companion document missing" in row["Sales Comments To Ops"]
    assert risk == "Medium"


def test_save_produces_two_sheets(tmp_path: Path):
    qr = QueryRegisterBuilder()
    extraction = _make_extraction("ADD")
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
    log_entry = _make_log_entry("Rejected", rejection_reason="missing signature")
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[])

    out_path = qr.save(tmp_path / "query_register.xlsx")
    assert out_path.exists()

    wb = load_workbook(out_path)
    # The query-log sheet is named for the run month, the way BIFM names
    # theirs ("August 2").
    month_sheet = date.today().strftime("%B %Y")
    assert wb.sheetnames == [month_sheet, "Recon"]

    query_log = wb[month_sheet]
    # Row 1 is the merged "Operations Use ONLY" banner; column headers are
    # on row 2, matching BIFM's own file.
    assert query_log.cell(row=2, column=1).value == "No."
    assert query_log.cell(row=2, column=3).value == "Type of Enquiry"
    assert query_log.cell(row=2, column=19).value == "Sales Comments To Ops"
    # The single rejected ADD row carries the form-code enquiry tag.
    assert query_log.cell(row=3, column=3).value == "Instruction-Additional"
    # Dates are real datetimes (the days-open formula subtracts them).
    assert isinstance(query_log.cell(row=3, column=2).value, datetime)
    # No. of Days Open (column R) is BIFM's live formula, not a frozen 0.
    days_open = query_log.cell(row=3, column=18).value
    assert isinstance(days_open, str) and days_open.startswith('=IF(AND(P3="Open"')

    recon = wb["Recon"]
    # Recon starts one column in (B) with no header over the label column,
    # and BIFM's exact header text in C2:E2.
    assert recon.cell(row=2, column=2).value is None
    assert recon.cell(row=2, column=3).value == "Number Of Instructions Received"
    assert recon.cell(row=2, column=4).value == "Processed"
    assert recon.cell(row=2, column=5).value == "Pending"
    for r in range(3, recon.max_row + 1):
        if recon.cell(row=r, column=2).value == "Additional Investments":
            assert recon.cell(row=r, column=3).value == 1
            # Pending is a live formula so hand-edited counts keep reconciling.
            assert recon.cell(row=r, column=5).value == f"=C{r}-D{r}"


def test_dropdown_validations_are_stored_without_a_leading_equals(tmp_path: Path):
    # Excel can silently strip data validations whose stored source range
    # carries a leading "=" ("we found a problem with some content"), which
    # presents to the user as the dropdowns being gone entirely. BIFM's own
    # file stores bare ranges ("$D$3:$D$17") - ours must too.
    qr = QueryRegisterBuilder()
    extraction = _make_extraction("ADD")
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
    log_entry = _make_log_entry("Rejected", rejection_reason="missing signature")
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[])
    out_path = qr.save(tmp_path / "query_register.xlsx")

    wb = load_workbook(out_path)
    ws = wb[date.today().strftime("%B %Y")]
    validations = ws.data_validations.dataValidation
    assert len(validations) == 4  # Type of Enquiry, Logged Via, Registered by, Status
    for dv in validations:
        assert dv.type == "list"
        assert not str(dv.formula1).startswith("="), f"formula1 {dv.formula1!r} would risk Excel stripping it"
        assert str(dv.formula1).startswith("$")


def _make_rich_extraction(form_code: str, **field_values) -> ExtractionResult:
    fields = {
        fid: FieldValue(field_id=fid, value=value, confidence=90.0)
        for fid, value in field_values.items()
    }
    return ExtractionResult(source_file="test.pdf", form_code=form_code, fields=fields)


def test_description_is_amount_plus_fund_in_bifm_style():
    # BIFM's own register writes "P25,000 MMF" - amount + fund shorthand -
    # in Query Description, not the validation problem.
    qr = QueryRegisterBuilder()
    extraction = _make_rich_extraction(
        "ADD", full_name="Theo Mabe",
        lump_sum_deposit_amount="25000",
        fund_name="BIFM Pula Money Market Fund",
    )
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
    log_entry = _make_log_entry("Rejected", rejection_reason="missing signature")
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[])

    row, _ = qr._query_rows[0]
    assert row["Query Description"] == "P25,000 MMF"
    assert "missing signature" in row["Sales Comments To Ops"]


def test_withdrawal_and_debit_descriptions():
    qr = QueryRegisterBuilder()
    log_entry = _make_log_entry("Rejected", rejection_reason="x")

    dis = _make_rich_extraction(
        "DIS", full_name="Nnelo", disinvestment_amount="1200",
        fund_name="BIFM Pula Money Market Fund",
    )
    qr.add_form(dis, ValidationReport(source_file="t", entity_number="E", results=[]), log_entry, [])

    debit = _make_rich_extraction(
        "DEBIT", full_name="Onalenna", sub_instruction_type="Cancel",
        fund_name="BIFM Balanced Prudential Fund",
    )
    qr.add_form(debit, ValidationReport(source_file="t", entity_number="E", results=[]), log_entry, [])

    assert qr._query_rows[0][0]["Query Description"] == "P1,200 MMF"
    assert qr._query_rows[1][0]["Query Description"] == "Debit order cancellation BPF"


def test_one_row_per_instruction_with_all_issues_combined():
    # A rejected instruction that ALSO raised flags gets ONE register row
    # (BIFM's No. column counts instructions), with every problem collected
    # into Sales Comments To Ops and the highest risk used for colouring.
    qr = QueryRegisterBuilder()
    extraction = _make_rich_extraction(
        "ADD", full_name="Theo Mabe",
        lump_sum_deposit_amount="5000", fund_name="BIFM Balanced Prudential Fund",
    )
    validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
    log_entry = _make_log_entry("Rejected", rejection_reason="wrong banking details")
    flag = PreValidationFlag(
        flag_id="kyc_missing", label="KYC companion document missing",
        triggered=True, rejection_risk="Medium", reason="No KYC form found in batch",
    )
    qr.add_form(extraction, validation, log_entry, prevalidation_flags=[flag])

    assert len(qr._query_rows) == 1
    row, risk = qr._query_rows[0]
    assert row["Query Description"] == "P5,000 BPF"
    assert "wrong banking details" in row["Sales Comments To Ops"]
    assert "KYC companion document missing" in row["Sales Comments To Ops"]
    assert risk == "High"


def test_add_form_is_thread_safe_under_concurrent_batches():
    # app.core.pipeline.run_batch shares one QueryRegisterBuilder across
    # multiple people's batches running concurrently - self._counter and
    # self._recon_counts are read-modify-write, so a missing lock would
    # silently lose increments (duplicate "No." values, undercounted
    # recon figures) under real concurrency, not just look wrong on paper.
    qr = QueryRegisterBuilder()
    n_calls = 50

    def make_and_add(i: int) -> None:
        extraction = _make_extraction("ADD", full_name=f"Investor {i}")
        validation = ValidationReport(source_file="test.pdf", entity_number="E1", results=[])
        log_entry = _make_log_entry("Rejected", rejection_reason="missing bank details")
        qr.add_form(extraction, validation, log_entry, prevalidation_flags=[])

    with ThreadPoolExecutor(max_workers=10) as pool:
        list(pool.map(make_and_add, range(n_calls)))

    assert len(qr._query_rows) == n_calls
    assert qr._counter == n_calls
    # No.s must be unique 1..n_calls - a lost increment would produce a
    # duplicate instead of every value appearing exactly once.
    numbers = sorted(row["No."] for row, _ in qr._query_rows)
    assert numbers == list(range(1, n_calls + 1))
    assert qr._recon_counts["Additional Investments"]["Received"] == n_calls
    assert qr._recon_counts["Additional Investments"]["Pending"] == n_calls
