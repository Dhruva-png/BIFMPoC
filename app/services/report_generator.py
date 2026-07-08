"""
Module 4: Excel Report Generator — produces the consolidated workbook with the
sheets specified by the client: Consolidated Investor Profile, Investor Master,
Beneficiary Details, Validation Flags, Pre-Validation Flags, Confidence Scores,
Processing Log.

Key metadata columns in Investor Master (from the requirements doc):
  Form Type, Fund Name, Fund Category, Investor Entity Number, Investor Full Name,
  Instruction Date (signed), Processing Cut-off, Instruction Status.

FAIL rows are red, WARNING rows are amber, PASS rows are green.

Investor Master now writes two physical rows per record: the data row,
followed immediately by a shaded/italic confidence sub-row showing the
extraction confidence for each field directly beneath its cell.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import (
    ExtractionResult,
    FieldValue,
    PreValidationFlag,
    ProcessingLogEntry,
    ValidationReport,
    ValidationStatus,
)
from app.utils.logger import get_logger
from config.settings import settings

logger = get_logger(__name__)

FILL_PASS    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_FAIL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

# NEW: styling for the confidence sub-row under each Investor Master record
CONFIDENCE_FONT = Font(italic=True, size=9, color="808080")
CONFIDENCE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

_STATUS_FILL = {
    ValidationStatus.PASS:    FILL_PASS,
    ValidationStatus.WARNING: FILL_WARNING,
    ValidationStatus.FAIL:    FILL_FAIL,
}

# Column order for the Investor Master sheet — mandatory metadata first,
# then extracted fields alphabetically.
_INVESTOR_MASTER_PRIORITY_COLS = [
    "source_file",
    "form_code",
    "form_type",           # derived: human-readable form name
    "entity_number",
    "full_name",
    "id_number",
    "id_type",
    "date_of_birth",
    "date_signed",
    "fund_name",
    "fund_number",
    "fund_category",       # derived: Money Market / Non-Money Market / GSGF
    "processing_cutoff",   # derived: 1:00 PM / 3:00 PM / Quarterly
    "overall_validation_status",
    "instruction_status",  # Submitted / Captured / Approved / Rejected
    "rejection_reason",
    "channel",             # Email / Walk-in
    "upload_date",
    "captured_by",
    "authorized_by",
    "contact_number",
    "email",
    "citizenship",
    "account_type",
]


# Column order for the Consolidated Investor Profile sheet - one row per
# batch/person, merging person-level fields picked up from any document in
# their batch (see app.services.consolidator).
_CONSOLIDATED_PROFILE_PRIORITY_COLS = [
    "person_key",
    "document_count",
    "source_documents",
    "entity_number",
    "entity_name",
    "full_name",
    "title",
    "id_number",
    "id_type",
    "id_expiry_date",
    "date_of_birth",
    "gender",
    "citizenship",
    "email",
    "contact_number",
    "residential_address",
    "postal_address",
    "occupation",
    "fund_name",
    "fund_number",
    "bank_account_name",
    "bank_name",
    "account_number",
    "branch_name",
    "branch_code",
    "account_type_banking",
    "authorized_signatory_name",
    "capacity",
]


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: list[str], min_width: int = 12, max_width: int = 45) -> None:
    for col_idx, header in enumerate(headers, start=1):
        longest = max(
            [len(str(header))] +
            [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, min_width), max_width)


def _ordered_investor_columns(all_rows: list[dict]) -> list[str]:
    """Return columns with priority cols first, then remaining alphabetically."""
    all_keys = {k for row in all_rows for k in row.keys()}
    priority = [c for c in _INVESTOR_MASTER_PRIORITY_COLS if c in all_keys]
    rest = sorted(all_keys - set(priority))
    return priority + rest


def build_single_document_workbook(
    extraction: ExtractionResult,
    validation: ValidationReport,
    log_entry: ProcessingLogEntry,
    prevalidation_flags: list[PreValidationFlag] | None = None,
) -> Workbook:
    """
    Builds a small, single-document workbook — NOT the multi-sheet batch
    report. One workbook per PDF, meant to sit right next to that PDF in
    its filed output folder (see app.core.pipeline._validate_and_file),
    so opening the folder shows a document and its own report side by
    side, with no cross-document/batch data mixed in.

    Two sheets:
      "Document"   — metadata (filename, form type, entity/name, status)
                     plus every extracted field with its value and
                     per-field confidence, colour-coded like the batch
                     report's Investor Master rows.
      "Validation" — this document's own validation + pre-validation
                     flag results (skipped if there are none).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Document"

    meta_rows = [
        ("Source file",          log_entry.original_filename),
        ("Filed as",             log_entry.new_filename),
        ("Form type",            log_entry.form_type_detected),
        ("Entity number",        validation.entity_number),
        ("Full name",            extraction.field_value("full_name") or ""),
        ("Classification conf.", f"{log_entry.classification_confidence:.0f}%"),
        ("Overall status",       validation.overall_status.value),
        ("Instruction status",   log_entry.instruction_status),
        ("Rejection reason",     log_entry.rejection_reason),
        ("Channel",              log_entry.channel),
        ("Processed at",         log_entry.timestamp),
    ]
    for label, value in meta_rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    status_fill = _STATUS_FILL.get(validation.overall_status)
    if status_fill:
        for col_idx in (1, 2):
            ws.cell(row=7, column=col_idx).fill = status_fill  # "Overall status" row

    ws.append([])
    _write_header(ws, ["Field", "Value", "Confidence", "Source"])
    ws.freeze_panes = None  # freeze_panes from _write_header targets row 1; not meaningful mid-sheet
    for fid, fv in extraction.fields.items():
        ws.append([fid, str(fv.value), f"{fv.confidence:.0f}%", fv.source])
        if fv.confidence >= 85:
            fill = FILL_PASS
        elif fv.confidence >= 65:
            fill = FILL_WARNING
        else:
            fill = FILL_FAIL
        for col_idx in range(1, 5):
            ws.cell(row=ws.max_row, column=col_idx).fill = fill
    _autosize(ws, ["Field", "Value", "Confidence", "Source"])
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 22)

    if extraction.beneficiaries:
        ws_b = wb.create_sheet("Beneficiaries")
        _write_header(ws_b, ["Name", "Relationship", "Split %"])
        for b in extraction.beneficiaries:
            ws_b.append([b.name, b.relationship, b.split_percent])
        _autosize(ws_b, ["Name", "Relationship", "Split %"])

    if validation.results or prevalidation_flags:
        ws_v = wb.create_sheet("Validation")
        headers = ["Field", "Status", "Message", "Value"]
        _write_header(ws_v, headers)
        for r in validation.results:
            ws_v.append([r.field_id, r.status.value, r.message, str(r.value or "")])
            fill = _STATUS_FILL.get(r.status)
            if fill:
                for col_idx in range(1, 5):
                    ws_v.cell(row=ws_v.max_row, column=col_idx).fill = fill
        if prevalidation_flags:
            ws_v.append([])
            ws_v.append(["Pre-Validation Flag", "Risk", "Triggered", "Reason"])
            for cell in ws_v[ws_v.max_row]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            for pf in prevalidation_flags:
                ws_v.append([pf.label, pf.rejection_risk, "YES" if pf.triggered else "no", pf.reason])
                if pf.triggered:
                    risk = str(pf.rejection_risk).lower()
                    fill = FILL_FAIL if risk.startswith("high") else FILL_WARNING
                    for col_idx in range(1, 5):
                        ws_v.cell(row=ws_v.max_row, column=col_idx).fill = fill
        _autosize(ws_v, headers)

    return wb


class ExcelReportBuilder:
    """
    Accumulates results across a batch run, then writes the consolidated workbook once.
    Call add_form(...) per processed document, then save().
    """

    def __init__(self) -> None:
        self._investor_rows: list[dict] = []
        self._investor_confidence: list[dict] = []  # parallel per-field confidence for each investor row
        self._beneficiary_rows: list[dict] = []
        self._validation_rows: list[tuple[dict, ValidationStatus]] = []
        self._log_entries: list[ProcessingLogEntry] = []
        self._consolidated_rows: list[dict] = []
        self._prevalidation_rows: list[tuple[dict, bool]] = []
        self._confidence_rows: list[dict] = []

    def merge_from(self, other: "ExcelReportBuilder") -> None:
        """
        Copies every row `other` has accumulated into this builder - used
        to build one combined ALL-PEOPLE workbook alongside each person's
        own per-person workbook, from the same already-completed
        extraction/validation results, without re-running the pipeline
        (and therefore without a second, possibly-different set of LLM
        calls) a second time just to populate a second report.
        """
        self._investor_rows.extend(other._investor_rows)
        self._investor_confidence.extend(other._investor_confidence)
        self._beneficiary_rows.extend(other._beneficiary_rows)
        self._validation_rows.extend(other._validation_rows)
        self._log_entries.extend(other._log_entries)
        self._consolidated_rows.extend(other._consolidated_rows)
        self._prevalidation_rows.extend(other._prevalidation_rows)
        self._confidence_rows.extend(other._confidence_rows)

    def add_form(
        self,
        extraction: ExtractionResult,
        validation: ValidationReport,
        log_entry: ProcessingLogEntry,
        prevalidation_flags: list[PreValidationFlag] | None = None,
    ) -> None:
        flat = extraction.to_flat_dict()
        flat["entity_number"] = validation.entity_number
        flat["overall_validation_status"] = validation.overall_status.value
        # Workflow/audit metadata (Metadata Fields for Filing table) lives on
        # the ProcessingLogEntry since it's determined at filing time, but the
        # client expects to see it alongside the extracted data in Investor
        # Master too, not buried only in the Processing Log sheet.
        flat["instruction_status"] = log_entry.instruction_status
        flat["rejection_reason"] = log_entry.rejection_reason
        flat["channel"] = log_entry.channel
        flat["upload_date"] = log_entry.upload_date
        flat["captured_by"] = log_entry.captured_by
        flat["authorized_by"] = log_entry.authorized_by
        self._investor_rows.append(flat)

        # Per-field confidence, keyed the same way as `flat`, so it can be
        # looked up column-for-column when writing the confidence sub-row.
        self._investor_confidence.append(
            {fid: fv.confidence for fid, fv in extraction.fields.items()}
        )

        for b in extraction.beneficiaries:
            self._beneficiary_rows.append({
                "entity_number":    validation.entity_number,
                "source_file":      extraction.source_file,
                "beneficiary_name": b.name,
                "relationship":     b.relationship,
                "split_percent":    b.split_percent,
            })

        for r in validation.results:
            self._validation_rows.append((
                {
                    "entity_number": validation.entity_number,
                    "source_file":   extraction.source_file,
                    "form_code":     extraction.form_code,
                    "field":         r.field_id,
                    "value":         r.value,
                    "status":        r.status.value,
                    "message":       r.message,
                },
                r.status,
            ))

        self._log_entries.append(log_entry)

        # Section 8: Pre-validation Flags (rejection-risk signals) — one
        # row per applicable flag, triggered or not, so Ops sees a full
        # checklist rather than only exceptions.
        for pf in (prevalidation_flags or []):
            self._prevalidation_rows.append((
                {
                    "entity_number":  validation.entity_number,
                    "source_file":    extraction.source_file,
                    "form_code":      extraction.form_code,
                    "flag":           pf.label,
                    "rejection_risk": pf.rejection_risk,
                    "triggered":      "YES" if pf.triggered else "no",
                    "reason":         pf.reason,
                },
                pf.triggered,
            ))

        # Section 7: Confidence Scores column group — per-field extraction
        # confidence, most useful on handwritten/low-quality scans.
        for fid, fv in extraction.fields.items():
            self._confidence_rows.append({
                "entity_number": validation.entity_number,
                "source_file":   extraction.source_file,
                "form_code":     extraction.form_code,
                "field":         fid,
                "value":         fv.value,
                "confidence":    fv.confidence,
                "source":        fv.source,
            })

    def add_consolidated_profile(
        self,
        profile: dict[str, FieldValue],
        person_key: str,
        source_files: list[str],
    ) -> None:
        """
        Records one row for the "Consolidated Investor Profile" sheet: the
        merged person-level fields (entity_number, full_name, contact
        details, banking details, etc.) picked up from ACROSS every
        document in this person's batch — not just whichever single form
        happened to have that field filled in. One call per batch/person.
        """
        flat = {fid: fv.value for fid, fv in profile.items()}
        flat["person_key"] = person_key
        flat["document_count"] = len(source_files)
        flat["source_documents"] = ", ".join(source_files)
        self._consolidated_rows.append(flat)

    def _build_workbook(self) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)

        self._write_consolidated_profile(wb)
        self._write_investor_master(wb)
        self._write_beneficiary_details(wb)
        self._write_validation_flags(wb)
        self._write_prevalidation_flags(wb)
        self._write_confidence_scores(wb)
        self._write_processing_log(wb)
        return wb

    def save(self, output_path: Path | None = None) -> Path:
        output_path = output_path or (settings.paths.output_dir / settings.excel_report_name)
        wb = self._build_workbook()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("Saved consolidated Excel report to %s", output_path)
        return output_path

    def _write_consolidated_profile(self, wb: Workbook) -> None:
        """
        One row per batch/person — the merged view an Ops reviewer actually
        wants: "who is this, and what do we know about them from everything
        they handed in", rather than having to cross-reference 4 separate
        Investor Master rows to piece together one person's contact and
        banking details.
        """
        ws = wb.create_sheet("Consolidated Investor Profile", 0)
        if not self._consolidated_rows:
            _write_header(ws, ["No data"])
            return
        all_keys = {k for row in self._consolidated_rows for k in row.keys()}
        priority = [c for c in _CONSOLIDATED_PROFILE_PRIORITY_COLS if c in all_keys]
        rest = sorted(all_keys - set(priority))
        headers = priority + rest
        _write_header(ws, headers)
        for row in self._consolidated_rows:
            ws.append([row.get(h, "") for h in headers])
        _autosize(ws, headers)

    def _write_investor_master(self, wb: Workbook) -> None:
        """
        Each record occupies two physical rows: the data row, followed
        immediately by a shaded/italic sub-row showing the extraction
        confidence for each field in the cell directly below it. Columns
        with no confidence value (derived/metadata columns like
        'instruction_status', 'fund_category', etc.) are left blank in the
        sub-row rather than showing a misleading 0 or N/A.
        """
        ws = wb.create_sheet("Investor Master")
        if not self._investor_rows:
            _write_header(ws, ["No data"])
            return

        headers = _ordered_investor_columns(self._investor_rows)
        _write_header(ws, headers)

        status_col_idx = headers.index("overall_validation_status") + 1 if "overall_validation_status" in headers else None

        for row, conf in zip(self._investor_rows, self._investor_confidence):
            data_row_idx = ws.max_row + 1
            ws.append([row.get(h, "") for h in headers])

            # Color-code the data row by overall validation status
            if status_col_idx:
                status_val = ws.cell(row=data_row_idx, column=status_col_idx).value
                fill = None
                if status_val == "PASS":
                    fill = FILL_PASS
                elif status_val == "WARNING":
                    fill = FILL_WARNING
                elif status_val in ("FAIL", "ERROR"):
                    fill = FILL_FAIL
                if fill:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=data_row_idx, column=col_idx).fill = fill

            # Confidence sub-row directly beneath the data row
            conf_values = []
            for h in headers:
                c = conf.get(h)
                if isinstance(c, (int, float)):
                    conf_values.append(f"{c:.0f}%")
                else:
                    conf_values.append("")
            ws.append(conf_values)
            conf_row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=conf_row_idx, column=col_idx)
                cell.font = CONFIDENCE_FONT
                if conf_values[col_idx - 1]:
                    cell.fill = CONFIDENCE_FILL
            ws.row_dimensions[conf_row_idx].height = 12

        _autosize(ws, headers)

    def _write_beneficiary_details(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Beneficiary Details")
        headers = ["entity_number", "source_file", "beneficiary_name", "relationship", "split_percent"]
        _write_header(ws, headers)
        for row in self._beneficiary_rows:
            ws.append([row.get(h, "") for h in headers])
        _autosize(ws, headers)

    def _write_validation_flags(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Validation Flags")
        headers = ["entity_number", "source_file", "form_code", "field", "value", "status", "message"]
        _write_header(ws, headers)
        for row, status in self._validation_rows:
            ws.append([row.get(h, "") for h in headers])
            fill = _STATUS_FILL.get(status)
            if fill:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = fill
        _autosize(ws, headers)

    def _write_prevalidation_flags(self, wb: Workbook) -> None:
        """Section 8: one row per applicable flag per document (full
        checklist, not just exceptions). Triggered rows are colour-coded by
        rejection risk so Ops can triage High first."""
        ws = wb.create_sheet("Pre-Validation Flags")
        headers = ["entity_number", "source_file", "form_code", "flag", "rejection_risk", "triggered", "reason"]
        _write_header(ws, headers)
        for row, triggered in self._prevalidation_rows:
            ws.append([row.get(h, "") for h in headers])
            if triggered:
                risk = str(row.get("rejection_risk", "")).lower()
                fill = FILL_FAIL if risk.startswith("high") else (
                    FILL_WARNING if risk.startswith("medium") else PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                    )
                )
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = fill
        _autosize(ws, headers)

    def _write_confidence_scores(self, wb: Workbook) -> None:
        """Section 7: Confidence Scores column group — per-field extraction
        confidence. Low-confidence fields (below the validation-rules
        'medium' threshold) are highlighted amber/red for reviewer attention,
        especially relevant for handwritten forms."""
        ws = wb.create_sheet("Confidence Scores")
        headers = ["entity_number", "source_file", "form_code", "field", "value", "confidence", "source"]
        _write_header(ws, headers)
        for row in self._confidence_rows:
            ws.append([row.get(h, "") for h in headers])
            confidence = row.get("confidence")
            if isinstance(confidence, (int, float)):
                fill = FILL_PASS if confidence >= 85 else (FILL_WARNING if confidence >= 65 else FILL_FAIL)
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = fill
        _autosize(ws, headers)

    def _write_processing_log(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Processing Log")
        headers = [
            "timestamp", "original_filename", "new_filename",
            "form_type_detected", "classification_confidence",
            "validation_status", "instruction_status", "rejection_reason",
            "channel", "upload_date", "destination_path",
            "captured_by", "authorized_by", "processor_id",
        ]
        _write_header(ws, headers)
        for entry in self._log_entries:
            ws.append([getattr(entry, h, "") for h in headers])
        _autosize(ws, headers)