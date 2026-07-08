"""
MarvelAI conversion layer.

MarvelAI cannot consume Python objects, JSON, or Excel files directly — it
only accepts data as a nested list of ``[field_name, field_value,
confidence]`` rows:

    [
        ["full_name", "Dhruva", 100],
        ["id_number", "224922314", 98],
        ["fund_name", "BIFM Pula Money Market Fund", 100],
    ]

`MarvelConverter` is the single place responsible for producing that shape,
regardless of where the underlying data comes from. It is deliberately kept
independent of the OCR extraction pipeline: it only depends on the public
`ExtractionResult` / `FieldValue` schema types and on `openpyxl` for reading
generated Excel reports. It does not modify, call into, or import from
app.ocr, app.core, app.llm, or app.validation.

Adding a new source in the future (a database row, a JSON payload, etc.)
means adding a new `from_<source>(...)` method to `MarvelConverter` that
returns the same `MarvelRow` list — the MarvelAI-facing interface never
needs to change, and no existing method is touched.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Union

from openpyxl import load_workbook

from app.models.schemas import ExtractionResult
from app.utils.logger import get_logger

logger = get_logger(__name__)

# A single MarvelAI row: [field_name, field_value, confidence].
# `field_value` is typed as `Any` because it mirrors `FieldValue.value: Any`
# in app.models.schemas - extracted values are not restricted to primitives.
# `confidence` stays a plain float in practice (FieldValue.confidence is
# `float`, 0-100) but is left broad enough to tolerate a stray `None` from
# an upstream source that hasn't set one, rather than raising at conversion
# time over a non-essential value.
MarvelRow = list[Union[str, Any]]


class MarvelConverter:
    """
    Converts application data into the nested-list format required by
    MarvelAI.

    This class is the only integration point between the existing
    application (OCR pipeline, Excel reports) and MarvelAI. Every public
    method returns the same shape:

        [
            ["field_name", "field_value", confidence],
            ...
        ]

    so that MarvelAI never needs to know, or care, which method produced
    the data it received.

    The class holds no state and is safe to instantiate freely, e.g.
    ``MarvelConverter().from_excel(path)``.
    """

    #: Row indices (1-based, as used by openpyxl) for the expected layout
    #: of a Marvel-ready processing report: field names on row 1, field
    #: values on row 2, confidence scores on row 3.
    _FIELD_NAME_ROW = 1
    _FIELD_VALUE_ROW = 2
    _CONFIDENCE_ROW = 3

    def from_extraction_result(self, result: ExtractionResult) -> list[MarvelRow]:
        """
        Convert an in-memory `ExtractionResult` into MarvelAI rows.

        Iterates over `result.fields` (a mapping of ``field_id ->
        FieldValue``) and produces one row per field, using the field's
        value and confidence exactly as extracted — no transformation or
        recalculation of confidence is performed here.

        Args:
            result: The `ExtractionResult` produced by the OCR pipeline
                for a single processed document.

        Returns:
            A list of ``[field_name, field_value, confidence]`` rows, one
            per entry in ``result.fields``, in iteration order.

        Example:
            >>> converter = MarvelConverter()
            >>> converter.from_extraction_result(result)
            [["full_name", "Dhruva", 100], ["id_number", "224922314", 99]]
        """
        rows: list[MarvelRow] = []
        for field_id, field_value in result.fields.items():
            rows.append(self._build_row(field_id, field_value.value, field_value.confidence))
        return rows

    def from_excel(self, path: Union[str, Path]) -> list[MarvelRow]:
        """
        Convert a generated processing report Excel file into MarvelAI rows.

        Expects a workbook whose active sheet is laid out as:
            Row 1: field names
            Row 2: field values
            Row 3: confidence scores

        Columns where the field name is empty are skipped, so ragged or
        sparsely populated reports do not produce blank/garbage rows.

        Args:
            path: Path to the ``.xlsx`` processing report.

        Returns:
            A list of ``[field_name, field_value, confidence]`` rows, one
            per non-empty column.

        Raises:
            FileNotFoundError: If ``path`` does not exist.
            ValueError: If the workbook has fewer than 3 rows.

        Example:
            >>> excel_to_nested_list("output/BIFM_UT_Processing_Report.xlsx")
            [["full_name", "Dhruva", 100], ["id_number", "224922314", 99]]
        """
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Marvel source workbook not found: {path}")

        workbook = load_workbook(filename=path, data_only=True, read_only=True)
        try:
            worksheet = workbook.active
            if worksheet.max_row < self._CONFIDENCE_ROW:
                raise ValueError(
                    f"Expected at least {self._CONFIDENCE_ROW} rows "
                    f"(field names, values, confidence) in {path}, "
                    f"found {worksheet.max_row}."
                )

            rows: list[MarvelRow] = []
            for column in worksheet.iter_cols(min_row=1, max_row=self._CONFIDENCE_ROW):
                field_name = column[self._FIELD_NAME_ROW - 1].value
                if field_name is None or str(field_name).strip() == "":
                    # Ignore empty columns, per spec.
                    continue

                field_value = column[self._FIELD_VALUE_ROW - 1].value
                confidence = column[self._CONFIDENCE_ROW - 1].value
                rows.append(self._build_row(str(field_name).strip(), field_value, confidence))

            return rows
        finally:
            workbook.close()

    @staticmethod
    def _build_row(field_name: str, value: object, confidence: object) -> MarvelRow:
        """
        Build a single MarvelAI row in the exact required shape.

        Centralizing row construction here means every `from_*` method
        produces an identical, consistent shape and avoids duplicating the
        ``[name, value, confidence]`` list literal throughout the class.

        Args:
            field_name: The field's name/identifier.
            value: The field's extracted or stored value.
            confidence: The field's confidence score.

        Returns:
            A single ``[field_name, field_value, confidence]`` row.
        """
        return [field_name, value, confidence]


def extraction_result_to_nested_list(result: ExtractionResult) -> list[MarvelRow]:
    """
    Convenience wrapper around ``MarvelConverter().from_extraction_result``.

    Args:
        result: The `ExtractionResult` produced by the OCR pipeline for a
            single processed document.

    Returns:
        A list of ``[field_name, field_value, confidence]`` rows, ready to
        hand to MarvelAI.
    """
    return MarvelConverter().from_extraction_result(result)


def excel_to_nested_list(path: Union[str, Path]) -> list[MarvelRow]:
    """
    Convenience wrapper around ``MarvelConverter().from_excel``.

    Args:
        path: Path to the ``.xlsx`` processing report.

    Returns:
        A list of ``[field_name, field_value, confidence]`` rows, ready to
        hand to MarvelAI.
    """
    return MarvelConverter().from_excel(path)