"""
Module 6: Query Register Automation (Section 4, Output 5 / Section 7 of the
understanding document).

Produces a workbook matching BIFM's own updated Query Register Excel file
EXACTLY, including a structural quirk that isn't obvious from just looking
at it: four of the visible columns (Type of Enquiry, Logged Via,
Registered by, Status) each have a second, header-less column sitting next
to them (D, H, J, Q respectively) that is NOT a data field - it's the
dropdown picklist source for the real column, referenced by Excel's own
"list" data validation. This module writes those lists once near the top
and wires up matching dropdowns, exactly like BIFM's file.

Sheet 1 - the query log, named for the run month the way BIFM names theirs
("August 2") - 21 physical columns:
  A No.                                  L Checked by
  B Date                                 M Resolution Progress
  C Type of Enquiry   (dropdown)         N Date Submitted to Ops / Resolved
  D   [picklist: enquiry types]          O Date Captured
  E Client Name                          P Status              (dropdown)
  F Query Description                    Q   [picklist: Resolved/Open]
  G Logged Via        (dropdown)         R No. of Days Open (live formula,
  H   [picklist: logged-via channels]        same one BIFM's file uses)
  I Registered by      (dropdown)        S Sales Comments To Ops  \
  J   [picklist: staff names]            T Ops Resolution          } merged
  K Assigned to                          U Ops Resolution Date    / "Operations
                                                                    Use ONLY"
                                                                    banner,
                                                                    row 1

Two Excel-compat details learned the hard way from BIFM's own file:
  - Data-validation source ranges are stored WITHOUT a leading "=" (e.g.
    "$D$3:$D$17"). Excel can silently strip validations whose stored
    formula carries the "=", which presents as "the dropdowns are gone"
    even though openpyxl and LibreOffice both accept either form.
  - Date columns hold real datetime values, not ISO strings - the
    days-open formula subtracts them, and text dates turn it into #VALUE!.

Sheet 2 - Recon: daily/periodic reconciliation tracker (Received /
Processed / Pending) by instruction type, columns B-E to match BIFM's own
layout, which starts one column in from the sheet edge.

This is populated automatically from each batch run instead of being
hand-maintained by the Sales/Contact Center team, per Section 4 Output 5:
"Automatically log queried, rejected, or incomplete instructions into the
Query Register format shared by BIFM ... reducing manual entry by the
Sales/Contact Center team and keeping the recon counts current."

A "query" line is logged for:
  - Any instruction that came back REJECTED from validation (Section 3,
    Step 5's rejection flow) - e.g. wrong banking details.
  - Any triggered pre-validation flag (Section 3, Step 2's missing/
    incomplete documentation flow, plus Section 8's other rejection-risk
    signals), whether or not the instruction ultimately validated PASS/
    WARNING - a triggered flag is itself something Sales needs to follow
    up with the client about.
Clean straight-through instructions with no rejection and no triggered
flags don't get a Query Log row at all - only the Recon counts move.

Every row this module writes gets Status = "Open" - matching BIFM's own
convention that new entries start open and Sales/Ops flips them to
"Resolved" by hand once actioned. Descriptive detail about WHY something
is open (e.g. "Awaiting corrected form from client") goes in Resolution
Progress, not Status - Status is strictly the 2-value Open/Resolved
picklist BIFM's file uses.
"""
from __future__ import annotations

import threading
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import ExtractionResult, PreValidationFlag, ProcessingLogEntry, ValidationReport
from app.utils.logger import get_logger
from config.settings import settings

logger = get_logger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin"),) * 4)

FILL_HIGH_RISK = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_MEDIUM_RISK = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# "Operations Use ONLY" banner - matches BIFM's own file: large, bold, red,
# merged across the 3 Ops-only columns in row 1.
OPS_BANNER_FONT = Font(bold=True, size=16, color="FF0000")
OPS_BANNER_FILL = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

DATE_FORMAT = "dd/mm/yyyy"

# How many rows past the current data to keep the dropdown validations
# active for, so Sales/Ops can keep adding rows by hand (in Excel
# directly) with the picklists still working, same as BIFM's own template
# pre-applying validation across ~1700 rows in advance.
VALIDATION_ROW_BUFFER = 500

# ---------------------------------------------------------------------------
# Column layout - position (1-indexed) : header text (None = picklist-only,
# no visible header, matching BIFM's own file exactly).
# ---------------------------------------------------------------------------
_COLUMNS: list[tuple[str, str | None]] = [
    ("No.", "No."),
    ("Date", "Date"),
    ("Type of Enquiry", "Type of Enquiry"),
    ("_enquiry_type_list", None),
    ("Client Name", "Client Name"),
    ("Query Description", "Query Description"),
    ("Logged Via", "Logged Via"),
    ("_logged_via_list", None),
    ("Registered by", "Registered by"),
    ("_registered_by_list", None),
    ("Assigned to", "Assigned to"),
    ("Checked by", "Checked by"),
    ("Resolution Progress", "Resolution Progress"),
    ("Date Submitted to Ops / Resolved", "Date Submitted to Ops / Resolved"),
    ("Date Captured", "Date Captured"),
    ("Status", "Status"),
    ("_status_list", None),
    ("No. of Days Open", "No. of Days Open"),
    ("Sales Comments To Ops", "Sales Comments To Ops"),
    ("Ops Resolution", "Ops Resolution"),
    ("Ops Resolution Date", "Ops Resolution Date"),
]
_COL_LETTERS = "ABCDEFGHIJKLMNOPQRSTU"
_COL_INDEX = {key: i + 1 for i, (key, _) in enumerate(_COLUMNS)}  # 1-indexed

_COLUMN_WIDTHS = {
    "A": 7, "B": 12, "C": 30, "D": 26, "E": 26, "F": 38, "G": 15, "H": 24,
    "I": 14, "J": 18, "K": 18, "L": 15, "M": 30, "N": 14, "O": 14, "P": 12,
    "Q": 14, "R": 20, "S": 28, "T": 32, "U": 22,
}

# ---------------------------------------------------------------------------
# Dropdown picklists - written once as static reference data (rows 3+ in
# their respective columns), exactly matching BIFM's own file. These are
# NOT per-row data - they're what the dropdown on the real column offers.
# ---------------------------------------------------------------------------
# Dropdown picklist values, taken verbatim (trailing whitespace trimmed) from
# BIFM's updated Query Register template's D/H/J/Q helper columns.
ENQUIRY_TYPE_OPTIONS = [
    "Instruction-Additional", "Instruction-Withdrawal", "Instruction-Debit Order",
    "Instruction-Static", "Instruction-Switch", "New Business - Original",
    "New Business - Copy", "KYC Submission", "Product Enquiry", "General Enquiry",
    "Investor Portal", "Statement/Balance request", "Complaint", "Smart Invest",
    "Instruction-Update Investor details",
]
LOGGED_VIA_OPTIONS = ["Email", "Walk-in", "Telephone", "Whatsapp", "Other"]
REGISTERED_BY_OPTIONS = [
    "Comfort", "Tefo", "Theo", "Onalenna", "Gaolatlhe", "Goiponyeone",
    "Brian", "Naomi", "Amolemo", "Nnnelo", "Kago",
]
# Same order as BIFM's file (Q3 = Resolved, Q4 = Open).
STATUS_OPTIONS = ["Resolved", "Open"]

# Columns holding real dates - written as datetime values (not ISO strings)
# so the days-open formula can subtract them.
_DATE_COLUMNS = {"Date", "Date Submitted to Ops / Resolved", "Date Captured", "Ops Resolution Date"}


def _as_datetime(value):
    """ISO date/datetime strings -> datetime; everything else unchanged."""
    if isinstance(value, (datetime, date)) or value in ("", None):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return value


def _days_open_formula(row: int) -> str:
    """The exact No.-of-Days-Open formula from BIFM's own file: still open
    and not yet captured -> days since logged; captured -> capture date
    minus logged date; otherwise -> submitted date minus logged date."""
    return (
        f'=IF(AND(P{row}="Open",O{row}=""),(TODAY()-B{row}),'
        f'IF(AND(O{row}<>"",P{row}<>""),O{row}-B{row},N{row}-B{row}))'
    )

RECON_HEADERS = ["Instruction Type", "Received", "Processed", "Pending"]

RECON_INSTRUCTION_TYPES = [
    "New Business",
    "Additional Investments",
    "Withdrawals",
    "Debit Orders",
    "Switch/Transfer",
    "Static",
]

# Maps this app's form codes (config/form_types.json) -> the Recon sheet's
# instruction-type buckets. DIS and DIS_GSG both roll up to "Withdrawals"
# for recon purposes even though they're tracked separately everywhere
# else in this app (different cut-offs / lock-ins - see form_types.json).
# KYC is a companion document (Section 8), not its own instruction line.
# SWITCH is included for forward-compatibility with Output 1's
# classification list even though it isn't one of the six sample form
# types extracted in this POC (Section 5).
_FORM_CODE_TO_RECON_TYPE: dict[str, str | None] = {
    "APPFORM": "New Business",
    "ADD": "Additional Investments",
    "DIS": "Withdrawals",
    "DIS_GSG": "Withdrawals",
    "DEBIT": "Debit Orders",
    "STATIC": "Static",
    "SWITCH": "Switch/Transfer",
    "KYC": None,
}


def recon_type_for_form_code(form_code: str) -> str | None:
    return _FORM_CODE_TO_RECON_TYPE.get(form_code)


# The updated Query Register's "Type of Enquiry" picklist is
# instruction-oriented (Instruction-Additional, New Business - Original,
# KYC Submission, ...) rather than the old rejection/query-category list, so
# the enquiry tag is chosen from the document's form type. The free-text
# Query Description column still carries the full rejection/flag detail
# regardless of which category tag is picked here.
_FORM_CODE_TO_ENQUIRY_TYPE: dict[str, str] = {
    "APPFORM": "New Business - Original",
    "ADD": "Instruction-Additional",
    "DIS": "Instruction-Withdrawal",
    "DIS_GSG": "Instruction-Withdrawal",
    "DEBIT": "Instruction-Debit Order",
    "STATIC": "Instruction-Static",
    "SWITCH": "Instruction-Switch",
    "KYC": "KYC Submission",
}


def _map_enquiry_type(form_code: str) -> str:
    """Maps the document's form code onto BIFM's updated Type of Enquiry
    picklist. Falls back to "General Enquiry" (itself a valid list option)
    for any code without a specific instruction category."""
    return _FORM_CODE_TO_ENQUIRY_TYPE.get(form_code, "General Enquiry")


class QueryRegisterBuilder:
    """
    Accumulates queried/rejected/incomplete instructions across a batch
    run, plus Received/Processed/Pending counts by instruction type, then
    writes the two-sheet workbook once. Mirrors the add_form(...) / save()
    pattern of app.services.report_generator.ExcelReportBuilder so the two
    can be filled in from the same call site in the pipeline.

    One instance is shared across every document in a run, including
    documents from DIFFERENT people's batches processed concurrently by
    app.core.pipeline.run_batch - unlike ExcelReportBuilder (whose state is
    plain list.append calls, atomic under the GIL), self._counter and
    self._recon_counts are read-modify-write, so add_form is guarded by
    self._lock to stop two threads' calls from racing and silently losing
    an increment (duplicate "No." values, undercounted recon figures).
    """

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._query_rows: list[tuple[dict, str]] = []  # (row, rejection_risk) for colour-coding
        self._counter = 0
        self._recon_counts: dict[str, dict[str, int]] = {
            t: {"Received": 0, "Processed": 0, "Pending": 0} for t in RECON_INSTRUCTION_TYPES
        }

    def add_form(
        self,
        extraction: ExtractionResult,
        validation: ValidationReport,
        log_entry: ProcessingLogEntry,
        prevalidation_flags: list[PreValidationFlag] | None = None,
    ) -> None:
        """One call per processed document - same call site as
        ExcelReportBuilder.add_form in app.core.pipeline."""
        with self._lock:
            self._add_form_locked(extraction, validation, log_entry, prevalidation_flags)

    def _add_form_locked(
        self,
        extraction: ExtractionResult,
        validation: ValidationReport,
        log_entry: ProcessingLogEntry,
        prevalidation_flags: list[PreValidationFlag] | None = None,
    ) -> None:
        recon_type = recon_type_for_form_code(extraction.form_code)
        if recon_type:
            counts = self._recon_counts[recon_type]
            counts["Received"] += 1
            if log_entry.instruction_status in ("Captured", "Approved"):
                counts["Processed"] += 1
            else:
                # Rejected or still Submitted - sits with Sales/Ops until
                # resolved, so it counts as Pending for recon purposes.
                counts["Pending"] += 1

        triggered = [f for f in (prevalidation_flags or []) if f.triggered]
        is_rejected = log_entry.instruction_status == "Rejected"

        if not is_rejected and not triggered:
            return  # clean straight-through instruction - nothing to log

        full_name = (
            extraction.field_value("full_name")
            or extraction.field_value("entity_number")
            or "Unknown"
        )
        today = date.today().isoformat()
        enquiry_type = _map_enquiry_type(extraction.form_code)

        if is_rejected:
            self._add_row(
                enquiry_type=enquiry_type,
                client_name=str(full_name),
                description=log_entry.rejection_reason or "Validation failed",
                resolution_progress="Pending correction from client",
                rejection_risk="High",
                today=today,
                log_entry=log_entry,
            )

        for flag in triggered:
            self._add_row(
                enquiry_type=enquiry_type,
                client_name=str(full_name),
                description=f"{flag.label}: {flag.reason}" if flag.reason else flag.label,
                resolution_progress="Pending follow-up with client",
                rejection_risk=flag.rejection_risk,
                today=today,
                log_entry=log_entry,
            )

    def _add_row(
        self, *, enquiry_type: str, client_name: str, description: str,
        resolution_progress: str, rejection_risk: str, today: str,
        log_entry: ProcessingLogEntry,
    ) -> None:
        self._counter += 1
        row = {
            "No.": self._counter,
            "Date": today,
            "Type of Enquiry": enquiry_type,
            "Client Name": client_name,
            "Query Description": description,
            "Logged Via": log_entry.channel,
            "Registered by": log_entry.captured_by,
            "Assigned to": "",
            "Checked by": "",
            "Resolution Progress": resolution_progress,
            "Date Submitted to Ops / Resolved": "",
            "Date Captured": log_entry.upload_date,
            "Status": "Open",
            "Sales Comments To Ops": "",
            "Ops Resolution": "",
            "Ops Resolution Date": "",
        }
        self._query_rows.append((row, rejection_risk))

    def save(self, output_path: Path | None = None) -> Path:
        output_path = output_path or (settings.paths.output_dir / settings.query_register_report_name)
        wb = Workbook()
        wb.remove(wb.active)
        self._write_query_log(wb)
        self._write_recon(wb)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("Saved Query Register to %s", output_path)
        return output_path

    # -----------------------------------------------------------------
    # Query Log sheet
    # -----------------------------------------------------------------

    def _write_query_log(self, wb: Workbook) -> None:
        # BIFM names the query-log sheet for the period it covers
        # ("August 2"); ours is named for the run month.
        ws = wb.create_sheet(date.today().strftime("%B %Y"))

        # Row 1: "Operations Use ONLY" banner, merged over the 3 Ops-only
        # columns (Sales Comments to Ops / Ops Resolution / Ops Resolution
        # Date - the last 3 columns), matching BIFM's file exactly.
        last3_start = len(_COLUMNS) - 2  # 1-indexed start of the last 3 columns
        start_letter = _COL_LETTERS[last3_start - 1]
        end_letter = _COL_LETTERS[len(_COLUMNS) - 1]
        ws.merge_cells(f"{start_letter}1:{end_letter}1")
        banner_cell = ws[f"{start_letter}1"]
        banner_cell.value = "Operations Use ONLY"
        banner_cell.font = OPS_BANNER_FONT
        banner_cell.fill = OPS_BANNER_FILL
        banner_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Row 2: column headers (blank for the picklist-only columns).
        for idx, (_key, header) in enumerate(_COLUMNS, start=1):
            letter = _COL_LETTERS[idx - 1]
            cell = ws.cell(row=2, column=idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 26

        # Static picklist source values - written once, not per-row data.
        self._write_picklist(ws, "D", ENQUIRY_TYPE_OPTIONS)
        self._write_picklist(ws, "H", LOGGED_VIA_OPTIONS)
        self._write_picklist(ws, "J", REGISTERED_BY_OPTIONS)
        self._write_picklist(ws, "Q", STATUS_OPTIONS)

        # Data rows, starting row 3.
        row_idx = 3
        for row, risk in self._query_rows:
            for key, _header in _COLUMNS:
                if key.startswith("_"):
                    continue  # picklist-only column, no per-row data
                col_idx = _COL_INDEX[key]
                if key == "No. of Days Open":
                    # Live formula, exactly as in BIFM's own file - not a
                    # frozen snapshot that goes stale the day after the run.
                    cell = ws.cell(row=row_idx, column=col_idx, value=_days_open_formula(row_idx))
                else:
                    value = _as_datetime(row.get(key, "")) if key in _DATE_COLUMNS else row.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if isinstance(value, (datetime, date)):
                        cell.number_format = DATE_FORMAT
                cell.border = THIN_BORDER

            risk_l = str(risk).lower()
            fill = FILL_HIGH_RISK if risk_l.startswith("high") else (
                FILL_MEDIUM_RISK if risk_l.startswith("medium") else None
            )
            if fill:
                for key, _header in _COLUMNS:
                    if key.startswith("_"):
                        continue
                    ws.cell(row=row_idx, column=_COL_INDEX[key]).fill = fill
            row_idx += 1

        last_data_row = max(row_idx - 1, 3)
        validation_last_row = last_data_row + VALIDATION_ROW_BUFFER

        # Pre-fill the days-open formula down the buffer rows too, exactly
        # like BIFM's file (theirs runs to row ~1750), so rows Sales adds
        # by hand in Excel compute without anyone copying the formula down.
        days_col = _COL_INDEX["No. of Days Open"]
        for r in range(row_idx, validation_last_row + 1):
            ws.cell(row=r, column=days_col, value=_days_open_formula(r))

        self._add_dropdown(ws, "C", f"$D$3:$D${2 + len(ENQUIRY_TYPE_OPTIONS)}", 3, validation_last_row)
        self._add_dropdown(ws, "G", f"$H$3:$H${2 + len(LOGGED_VIA_OPTIONS)}", 3, validation_last_row)
        self._add_dropdown(ws, "I", f"$J$3:$J${2 + len(REGISTERED_BY_OPTIONS)}", 3, validation_last_row)
        self._add_dropdown(ws, "P", f"$Q$3:$Q${2 + len(STATUS_OPTIONS)}", 3, validation_last_row)

        for letter, width in _COLUMN_WIDTHS.items():
            ws.column_dimensions[letter].width = width

        # Matches BIFM's updated file: the header rows plus the first few
        # identifying columns (No. through Client Name) stay visible while
        # scrolling right into the "Operations Use ONLY" section.
        ws.freeze_panes = "F3"

        if not self._query_rows:
            logger.info("Query Log: no queries, rejections, or missing-documentation flags this run.")

    @staticmethod
    def _write_picklist(ws: Worksheet, column_letter: str, options: list[str]) -> None:
        for i, option in enumerate(options):
            ws[f"{column_letter}{3 + i}"] = option

    @staticmethod
    def _add_dropdown(ws: Worksheet, column_letter: str, source_range: str, first_row: int, last_row: int) -> None:
        # NO leading "=" on formula1: openpyxl writes it verbatim into the
        # sheet XML, and while LibreOffice tolerates "=$D$3:$D$17", real
        # Excel can flag it as invalid content and silently strip the
        # validation on open - i.e. the dropdowns vanish. BIFM's own file
        # stores the bare range ("$D$3:$D$16"), so we do exactly that.
        dv = DataValidation(type="list", formula1=source_range, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{column_letter}{first_row}:{column_letter}{last_row}")

    # -----------------------------------------------------------------
    # Recon sheet
    # -----------------------------------------------------------------

    def _write_recon(self, wb: Workbook) -> None:
        """Matches BIFM's own Recon layout exactly: starts one column in
        (B, not A), no header over the label column, header text
        "Number Of Instructions Received" / "Processed" / "Pending" in
        C2:E2, and live formulas for Pending and the TOTAL row so counts
        Ops edits by hand keep reconciling."""
        ws = wb.create_sheet("Recon")

        for col_idx, header in ((3, "Number Of Instructions Received"), (4, "Processed"), (5, "Pending")):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        row_idx = 3
        for instr_type in RECON_INSTRUCTION_TYPES:
            counts = self._recon_counts[instr_type]
            label_cell = ws.cell(row=row_idx, column=2, value=instr_type)
            label_cell.font = Font(bold=True)
            label_cell.border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=counts["Received"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=4, value=counts["Processed"]).border = THIN_BORDER
            pending = ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
            pending.border = THIN_BORDER
            row_idx += 1

        total_cell = ws.cell(row=row_idx, column=2, value="TOTAL")
        total_cell.font = Font(bold=True)
        for col_letter, col_idx in (("C", 3), ("D", 4), ("E", 5)):
            cell = ws.cell(row=row_idx, column=col_idx, value=f"=SUM({col_letter}3:{col_letter}{row_idx - 1})")
            cell.font = Font(bold=True)

        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 30
        for letter in ("D", "E"):
            ws.column_dimensions[letter].width = 14